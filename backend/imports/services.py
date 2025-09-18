#imports/services.py
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Iterable, Tuple, Dict, Any, Optional, List, Set
from datetime import time, datetime, timedelta
import re
import uuid

import pandas as pd
from django.apps import apps as django_apps
from django.db import connections, transaction, models
from django.core.exceptions import ValidationError
from django.utils import timezone
from eoi.models import EoiApp
from users.models import User, Campus
from units.models import Unit, UnitCourse, Course
from timetable.models import MasterClassTime, TimeTable, TimetableImportLog
    
try:
    from openpyxl import load_workbook
except Exception as e:
    load_workbook = None

# =========================
# Header validation helpers
# =========================
FALLBACK_EOI_TABLE = "eoi_imports"

REQUIRED_COLUMNS: Dict[str, set[str]] = {
    "eoi": {
        "tutor_email",
        "unit_code",
        "preference",
        "campus",
        "qualifications",
        "availability",
    },
    "master_classes": {
        "subject_code",
        "subject_description",
        "activity_group_code",
        "activity_code",
        "activity_description",
        "campus",
        "location",
        "day_of_week",
        "start_time",
        "duration",
        "weeks",
        "staff",
        "size",
    },
    "tutorial_allocations": {
        "unit_code",
        "day_of_week",
        "start_time",
        "end_time",
        "room",
        "tutor_email",
    },
}

# Flexible synonyms so slightly different headers still map correctly
HEADER_SYNONYMS: Dict[str, set[str]] = {
    # generic
    "tutor_email": {"TutorEmail", "Email", "Applicant", "ApplicantEmail", "applicant", "applicant_email"},
    "unit_code": {"UnitCode", "Unit", "Unit Code", "subject_code"},
    "preference": {"Preference", "Rank", "Priority"},
    "campus": {"Campus", "Location"},
    "qualifications": {"Qualifications", "Notes", "Experience"},
    "availability": {"Availability", "Available", "Times"},
    # master classes
    "subject_description": {"SubjectDescription", "Subject Name", "Unit Name"},
    "activity_group_code": {"ActivityGroup", "GroupCode", "Group"},
    "activity_code": {"ActivityCode", "Activity"},
    "activity_description": {"ActivityDescription", "Activity Desc"},
    "location": {"Location", "RoomLocation"},
    "day_of_week": {"Day", "DayOfWeek"},
    "start_time": {"Start", "StartTime"},
    "duration": {
        "Duration", "duration", "Duration (minutes)", "Duration(min)",
        "Duration mins", "duration (mins)", "duration (min)", "Minutes"
    },
    "weeks": {"Weeks", "TeachingWeeks"},
    "staff": {"Staff", "StaffName", "Lecturer"},
    "size": {"Size", "Capacity"},
    "end_time": {"End", "EndTime"},
    "room": {"Room", "Location"},
}

_EOI_SYNONYMS = {
    "unit_code": {"unit_code", "unit", "unitcode", "code"},
    "tutor_email": {"tutor_email", "email", "email_address", "tutoremail"},
    "preference": {"preference", "pref", "rank", "order"},
    "campus": {"campus", "campus_name", "location"},
    "qualifications": {"qualifications", "skills", "technical_skills", "why"},
    "availability": {"availability", "hours", "tutoring_hours"},
    "gpa": {"gpa"},
    "references": {"references", "supervisor"},
}

_CAMPUS_ALIAS = {
    "hobart": "SB",
    "sandy bay": "SB",
    "launceston": "IR",
    "inveresk": "IR",
    "online": "ONLINE",
    "distance": "ONLINE",
}

UNIT_CODE_RX = re.compile(r"\b([A-Z]{3}\d{3})\b")

EMAIL_KEYS = {"email", "email address", "email address*"}
AVAIL_KEYS = {
    "availability",
    "total number of tutoring hours you wish to work - please consider your scholarship / visa condition",
    "total number of tutoring hours you wish to work",
    "total number of tutoring hours",
    "hours",
}
QUAL_KEYS = {
    "what technical and/or other skills do you have for your preferences? and why do you want to teach this unit?",
    "what technical and/or other skills do you have for your preferences? and why?",
    "what technical and/or other skills",
    "skills",
    "why do you want to teach this unit?",
}

CAMPUS_NAMES = {"hobart", "launceston", "online"}

# campus synonym map
CAMPUS_MAP = {
    "sb": "SB", "sandy bay": "SB", "hobart": "SB",
    "ir": "IR", "inveresk": "IR", "launceston": "IR",
    "online": "ONLINE", "distance": "ONLINE", "web": "ONLINE",
}

def _norm_campus(value: str) -> str:
    v = (value or "").strip().lower()
    return CAMPUS_MAP.get(v, (value or "").strip().upper() or "SB")

def _slot_key(uc_id: int, day_name: str, start_time: time, room: str) -> tuple:
    """Canonical identity for a timetable row."""
    return (
        uc_id,
        (day_name or "").strip(),
        (start_time.strftime("%H:%M:%S") if isinstance(start_time, time) else str(start_time)),
        (room or "").strip().upper(),
    )

def _infer_term_year_from_db(using: str) -> tuple[str, int]:
    name = connections[using].settings_dict.get("NAME", "") or using
    y = re.search(r"(\d{4})", name)
    t = re.search(r"(S\d+|T\d+)", name, re.I)
    year = int(y.group(1)) if y else timezone.now().year
    term = t.group(1).upper() if t else "S1"
    return term, year

def _get_default_course(using: str) -> Course:
    """
    UnitCourse.course is NOT NULL in DB.
    We attach all UnitCourse rows to a placeholder Course ('GEN') unless you later wire a real course.
    """
    course, _ = Course.objects.using(using).get_or_create(
        course_code="GEN",
        defaults={"course_name": "General (placeholder)"}
    )
    return course

def _resolve_campus(using: str, raw: str) -> Optional[Campus]:
    key = (raw or "").strip().lower()
    key = CAMPUS_MAP.get(key, CAMPUS_MAP.get(key.replace(".", ""), None))
    if not key:
        return None
    return Campus.objects.using(using).filter(campus_name__iexact=key).first()

# Flexible header matching (lowercased, spaces collapsed)
def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()

def _norm_name(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", s.strip().lower())

def _best_mapping(model) -> Optional[dict]:
    """
    Try to map model fields to our canonical keys using synonyms.
    Returns a mapping like {"unit_code": "unit", "tutor_email": "email", ...}
    or None if the model can't support the required data.
    """
    # collect concrete field names on the model
    all_fields = []
    for f in model._meta.get_fields():
        # only concrete, non-M2M fields with a column
        if getattr(f, "concrete", False) and getattr(f, "attname", None) and not getattr(f, "many_to_many", False):
            all_fields.append(f.attname)

    norm_fields = {_norm_name(n): n for n in all_fields}  # norm->actual

    mapping = {}
    missing_hard = []
    for canon_key, syns in _EOI_SYNONYMS.items():
        found = None
        for syn in syns:
            if _norm_name(syn) in norm_fields:
                found = norm_fields[_norm_name(syn)]
                break
        if found:
            mapping[canon_key] = found
        else:
            # We consider unit_code, tutor_email, campus as hard requirements for a model target
            if canon_key in {"unit_code", "tutor_email", "campus"}:
                missing_hard.append(canon_key)
    if missing_hard:
        return None

    # soft fields get defaults later if absent
    return mapping

def _find_eoi_destination(using: str):
    try:
        app_cfg = django_apps.get_app_config("eoi")
    except LookupError:
        return None, None

    for model in app_cfg.get_models():
        if not _model_has_auto_pk(model):
            continue  # avoid tables that will demand explicit PK values
        mapping = _best_mapping(model)
        if mapping:
            return model, mapping
    return None, None

# --- ensure/fix master_eoi DDL ----------------------------------------------

def _existing_columns(using: str, table: str) -> list[tuple]:
    # returns DESCRIBE rows: (Field, Type, Null, Key, Default, Extra)
    with connections[using].cursor() as c:
        c.execute(f"SHOW COLUMNS FROM `{table}`;")
        return list(c.fetchall())

def _ensure_fallback_eoi_table(using: str):
    ddl = f"""
    CREATE TABLE IF NOT EXISTS `{FALLBACK_EOI_TABLE}` (
        `id` INT NOT NULL AUTO_INCREMENT,
        `unit_code` VARCHAR(32) NOT NULL,
        `tutor_email` VARCHAR(255) NOT NULL,
        `preference` INT NOT NULL DEFAULT 0,
        `campus` VARCHAR(32) NOT NULL,
        `qualifications` TEXT NULL,
        `availability` TEXT NULL,
        PRIMARY KEY (`id`)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
    """
    with connections[using].cursor() as c:
        c.execute(ddl)

def _ensure_master_eoi_table(using: str):
    ddl = """
    CREATE TABLE IF NOT EXISTS `master_eoi` (
      `id` INT NOT NULL AUTO_INCREMENT,
      `unit_code` VARCHAR(32) NOT NULL,
      `tutor_email` VARCHAR(255) NOT NULL,
      `preference` INT DEFAULT 0,
      `campus` VARCHAR(64) NOT NULL,
      `qualifications` LONGTEXT,
      `availability` VARCHAR(128),
      PRIMARY KEY (`id`),
      KEY `idx_eoi_unit` (`unit_code`),
      KEY `idx_eoi_email` (`tutor_email`)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
    """
    # DDL must NOT be inside an atomic block
    with connections[using].cursor() as c:
        c.execute(ddl)

def _ensure_master_eoi_columns(using: str):
    # add any missing non-PK columns on older schemas
    wanted = {
        "unit_code": "VARCHAR(32) NOT NULL",
        "tutor_email": "VARCHAR(255) NOT NULL",
        "preference": "INT DEFAULT 0",
        "campus": "VARCHAR(64) NOT NULL",
        "qualifications": "LONGTEXT",
        "availability": "VARCHAR(128)",
    }
    existing = {row[0] for row in _existing_columns(using, "master_eoi")}
    missing = [col for col in wanted if col not in existing]
    if not missing:
        return
    with connections[using].cursor() as c:
        for col in missing:
            c.execute(f"ALTER TABLE `master_eoi` ADD COLUMN `{col}` {wanted[col]};")

def _ensure_master_eoi_pk_auto(using: str):
    """
    Ensure the primary key column is AUTO_INCREMENT, even on legacy schemas
    where the PK is named `master_eoi_id` (or something else) and lacks AUTO_INCREMENT.
    """
    rows = _existing_columns(using, "master_eoi")
    # rows: (Field, Type, Null, Key, Default, Extra)
    pk_name = None
    pk_extra = ""
    for f, t, n, k, d, x in rows:
        if k == "PRI":
            pk_name, pk_extra = f, x or ""
            break

    with connections[using].cursor() as c:
        if not pk_name:
            # no PK at all -> add an auto id
            c.execute("ALTER TABLE `master_eoi` ADD COLUMN `id` INT NOT NULL AUTO_INCREMENT PRIMARY KEY;")
            return

        if "auto_increment" not in pk_extra.lower():
            # make existing PK auto-increment (assume integer-like)
            c.execute(f"ALTER TABLE `master_eoi` MODIFY COLUMN `{pk_name}` INT NOT NULL AUTO_INCREMENT;")

def _ensure_baseline_campuses(*, using: str) -> dict[str, Campus]:
    """
    Make sure SB(id=2), IR(id=1), ONLINE(id=3) exist.
    (IDs chosen to match your DB; if an ID is taken, we keep the row by name.)
    Returns a dict {"SB": Campus, "IR": Campus, "ONLINE": Campus}.
    """
    want = [
        {"id": 1, "campus_name": "IR",     "campus_location": "Launceston"},
        {"id": 2, "campus_name": "SB",     "campus_location": "Hobart"},
        {"id": 3, "campus_name": "ONLINE", "campus_location": "Online"},
    ]
    got = {}
    with transaction.atomic(using=using):
        for row in want:
            obj = Campus.objects.using(using).filter(campus_name=row["campus_name"]).first()
            if obj is None:
                # try to create with explicit id if it's free
                try:
                    obj = Campus(id=row["id"], campus_name=row["campus_name"], campus_location=row["campus_location"])
                    obj.save(using=using, force_insert=True)
                except Exception:
                    # fallback without forcing the id
                    obj, _ = Campus.objects.using(using).get_or_create(
                        campus_name=row["campus_name"],
                        defaults={"campus_location": row["campus_location"]}
                    )
            got[row["campus_name"]] = obj
    return got

def _find_unit_code(ws) -> Optional[str]:
    # Prefer sheet title first
    m = UNIT_CODE_RX.search(ws.title or "")
    if m:
        return m.group(1)
    # Then search first 8 rows
    for r in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        for v in r:
            m = UNIT_CODE_RX.search(str(v or ""))
            if m:
                return m.group(1)
    return None

def _detect_campus_row(values):
    """
    Accepts rows that contain the campus name ANYWHERE on the row,
    including banner rows like 'Select Tutor with preference number HOBART'.
    Returns the pretty campus token ('Hobart','Launceston','Online') or None.
    """
    tokens = [str(v or "").strip() for v in values if str(v or "").strip()]
    if not tokens:
        return None

    # Exact token match first
    for t in tokens:
        n = _norm(t)
        if n in CAMPUS_NAMES:
            return t.strip().title()

    # Fuzzy: campus embedded in a longer sentence
    joined = " ".join(tokens).lower()
    for name in CAMPUS_NAMES:
        if name in joined:
            return name.title()

    return None

def _find_header(ws, start_row: int) -> Optional[Dict[int, str]]:
    """
    Scan down from start_row to find the row that contains 'Email' etc.
    Some templates have one or two preheader rows; look a bit farther.
    """
    # look up to 12 rows below the campus banner
    for r in range(start_row, min(ws.max_row, start_row + 12)):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        normed = [_norm(v) for v in row]
        if any(k in normed for k in (_norm("Email Address"), "email", "email address*")):
            return {c: normed[c - 1] for c in range(1, len(normed) + 1)}
    return None

def _col_index(header_map: Dict[int, str], keys: Iterable[str]) -> Optional[int]:
    target = {_norm(k) for k in keys}
    for c, name in header_map.items():
        if name in target:
            return c
    # Heuristic partial matches
    for c, name in header_map.items():
        if any(_norm(k) in name for k in keys):
            return c
    return None

def _iter_normalized_rows(ws) -> Iterable[Dict[str, Any]]:
    """
    Yields dicts with keys:
      unit_code, campus, tutor_email, preference, availability, qualifications
    """
    unit_code = _find_unit_code(ws)
    current_campus = None
    r = 1
    while r <= ws.max_row:
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        campus_here = _detect_campus_row(row_vals)
        if campus_here:
            current_campus = campus_here
            # Find the table header beneath the campus banner
            header = _find_header(ws, r + 1)
            if not header:
                r += 1
                continue
            c_email = _col_index(header, EMAIL_KEYS)
            c_avail = _col_index(header, AVAIL_KEYS)
            c_qual  = _col_index(header, QUAL_KEYS)

            # Guess a preference column:
            # the first column to the left of "Name"/"Email" with small integers often holds the rank
            possible_pref_cols = [1]
            # find "name" column to try col-1 as pref
            for c, name in header.items():
                if "name" in name:
                    if c > 1:
                        possible_pref_cols.insert(0, c - 1)
            # scan down until the next blank email streak
            rr = r + 1
            while rr <= ws.max_row:
                vals = [ws.cell(rr, c).value for c in range(1, ws.max_column + 1)]
                email = str(vals[c_email - 1]).strip() if c_email else ""
                if not email or "@" not in email:
                    # assume table ended once we hit two empty-email rows in a row
                    # (next campus banner will reset things anyway)
                    # check next row quickly
                    nxt = [ws.cell(rr + 1, c).value for c in range(1, ws.max_column + 1)] if rr + 1 <= ws.max_row else []
                    if not nxt or not any("@" in str(v or "") for v in nxt):
                        break
                    rr += 1
                    continue

                # preference
                pref = None
                for pc in possible_pref_cols:
                    if 1 <= pc <= len(vals):
                        try:
                            x = vals[pc - 1]
                            if x is not None and str(x).strip() != "":
                                pref_num = int(str(x).strip())
                                if 0 <= pref_num <= 99:
                                    pref = pref_num
                                    break
                        except Exception:
                            pass
                # fallback: implicit order
                if pref is None:
                    pref = 0

                availability = (str(vals[c_avail - 1]).strip() if c_avail else "") or ""
                qualifications = (str(vals[c_qual - 1]).strip() if c_qual else "") or ""

                yield {
                    "unit_code": unit_code or "",
                    "campus": current_campus,
                    "tutor_email": email.lower(),
                    "preference": pref,
                    "availability": availability,
                    "qualifications": qualifications,
                }
                rr += 1
            # continue scanning after this table
            r = rr
            continue

        r += 1

def _normalize_workbook_to_rows(fobj) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise ValidationError("openpyxl is not installed on the server.")
    wb = load_workbook(fobj, data_only=True, read_only=True)
    rows: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        for rec in _iter_normalized_rows(ws):
            # skip obviously broken rows
            if not rec.get("tutor_email") or "@" not in rec["tutor_email"]:
                continue
            if not rec.get("unit_code"):
                continue
            rows.append(rec)
    if not rows:
        raise ValidationError(
            "Could not find any EOI rows. Ensure each unit tab contains a campus section "
            "with an 'Email Address' column. Hidden columns are supported."
        )
    return rows

def _find_eoi_model() -> Optional[Any]:
    try:
        app_cfg = django_apps.get_app_config("eoi")
    except LookupError:
        return None
    needed = {"unit_code", "tutor_email", "preference", "campus", "qualifications", "availability"}
    for model in app_cfg.get_models():
        fields = {f.name for f in model._meta.get_fields() if getattr(f, "attname", None)}
        if needed.issubset(fields):
            return model
    return None

def _table_has_columns(using: str, table: str, cols: Iterable[str]) -> bool:
    cols = list(cols)
    with connections[using].cursor() as c:
        try:
            c.execute(f"SHOW COLUMNS FROM `{table}`;")
            have = {row[0] for row in c.fetchall()}
            return set(cols).issubset(have)
        except Exception:
            return False

def _strip(s) -> str:
    return str(s).strip() if s is not None else ""


def _normalise_headers(df: pd.DataFrame) -> Dict[str, str]:
    """
    Returns a mapping {wanted_key -> actual_column_name_in_df}
    """
    mapping: Dict[str, str] = {}
    dfcols = [str(c).strip() for c in df.columns]
    lower_to_actual = {c.lower(): c for c in dfcols}

    for want in set().union(*REQUIRED_COLUMNS.values()) | set(HEADER_SYNONYMS.keys()):
        # exact lowercase match
        if want.lower() in lower_to_actual:
            mapping[want] = lower_to_actual[want.lower()]
            continue
        # synonyms
        for syn in HEADER_SYNONYMS.get(want, set()):
            if syn in dfcols:
                mapping[want] = syn
                break
    return mapping


def _validate_headers(kind: str, df: pd.DataFrame) -> Dict[str, str]:
    mapping = _normalise_headers(df)
    missing = [c for c in REQUIRED_COLUMNS[kind] if c not in mapping]
    if missing:
        raise ValueError(
            f"Spreadsheet is missing required columns for '{kind}': {', '.join(missing)}"
        )
    return mapping


# ============
# Conversions
# ============

DAY_ALIASES = {
    "mon": "Monday",
    "monday": "Monday",
    "tue": "Tuesday",
    "tues": "Tuesday",
    "tuesday": "Tuesday",
    "wed": "Wednesday",
    "wednesday": "Wednesday",
    "thu": "Thursday",
    "thur": "Thursday",
    "thurs": "Thursday",
    "thursday": "Thursday",
    "fri": "Friday",
    "friday": "Friday",
    "sat": "Saturday",
    "saturday": "Saturday",
    "sun": "Sunday",
    "sunday": "Sunday",
}


def _as_day_name(s: str) -> str:
    if not s:
        return s
    key = s.strip().lower()
    return DAY_ALIASES.get(key, s.strip())


def _as_time(cell) -> str:
    """
    Return 'HH:MM:SS' string; supports Excel/strings/pandas time.
    """
    return _to_time(cell)

def _end_time_from_start_and_duration(start, duration_min: int):
    if not start or not isinstance(duration_min, int) or duration_min <= 0:
        return None
    dt = datetime(2000, 1, 1, start.hour, start.minute)
    dt2 = dt + timedelta(minutes=duration_min)
    return dt2.time()

def _canon_weeks_str(raw: str) -> str:
    """
    Accept values like '1-13', '1,2,3,5-7', even Excel numbers.
    Return a canonical string; default to '1-13' if blank.
    """
    s = (raw or "").strip()
    if not s:
        return "1-13"
    return s.replace(" ", "")

def _count_weeks(weeks: str) -> int:
    """
    Count distinct week numbers in a string like '1-13' or '1,2,3,5-7'.
    Defaults to 13 if blank.
    """
    s = (weeks or "").replace(" ", "")
    if not s:
        return 13
    seen = set()
    for tok in s.split(","):
        if not tok:
            continue
        if "-" in tok:
            a, b = tok.split("-", 1)
            try:
                x, y = int(a), int(b)
                if x > y:
                    x, y = y, x
                for w in range(max(1, x), min(52, y) + 1):
                    seen.add(w)
            except ValueError:
                continue
        else:
            try:
                w = int(tok)
                if 1 <= w <= 52:
                    seen.add(w)
            except ValueError:
                continue
    return max(1, len(seen))

def _model_has_auto_pk(model) -> bool:
    pk = model._meta.pk
    return isinstance(pk, (models.AutoField, models.BigAutoField, models.SmallAutoField))
# ==================
# Logging to control
# ==================

@dataclass
class ImportStats:
    ok: int = 0
    err: int = 0
    errors: list[str] = field(default_factory=list)

    def __post_init__(self):
        if self.errors is None:
            self.errors = []

    def log(self, msg: str):
        self.errors.append(msg)
        self.err += 1

    def inc(self):
        self.ok += 1


def _create_log(filename: str, uploaded_by: User) -> TimetableImportLog:
    # Keep logs on default DB
    return TimetableImportLog.objects.create(
        import_id=uuid.uuid4(),
        filename=filename,
        status="Processing",
        total_rows=0,
        processed_rows=0,
        error_rows=0,
        error_log="",
        uploaded_by=uploaded_by,
        created_at=timezone.now(),
    )


def _finalise_log(log: TimetableImportLog, stats: ImportStats, total: int, ok_status: str):
    log.total_rows = total
    log.processed_rows = stats.ok
    log.error_rows = stats.err
    log.error_log = "\n".join(stats.errors)
    log.status = "Completed" if stats.err == 0 else "Failed"
    log.completed_at = timezone.now()
    log.save(update_fields=[
        "total_rows", "processed_rows", "error_rows", "error_log", "status", "completed_at"
    ])


# ===========================
# Master class times (MCT) import
# ===========================
def import_master_classes_xlsx(file_like, job, using: str):
    """
    Import Master Class List.xlsx into:
      - units (unit_code/unit_name)
      - unit_courses (unit + course + campus + term/year)
      - master_class_time (template rows)
      - timetable (actual sessions used by allocation/clash)
    Guarantees:
      - campus exists for SB/IR/ONLINE
      - unit_courses.course_id is NEVER NULL
      - timetable.teaching_weeks is ALWAYS set
    """
    df = pd.read_excel(file_like)
    col = _validate_headers("master_classes", df)

    # Ensure campuses baseline
    campuses = _ensure_baseline_campuses(using=using)

    term, year = _infer_term_year_from_db(using)
    course = _get_default_course(using)

    stats = ImportStats()
    log = _create_log(getattr(job.file, "name", "master_classes.xlsx"), job.created_by)
    total = len(df)
    seen_slots: set[tuple] = set()

    for i, row in df.iterrows():
        try:
            subject_code        = _strip(row[col["subject_code"]])
            subject_description = _strip(row[col["subject_description"]])
            activity_group_code = _strip(row[col["activity_group_code"]])
            activity_code       = _strip(row[col["activity_code"]])
            activity_description= _strip(row[col["activity_description"]])
            campus_txt          = _strip(row[col["campus"]])
            campus_code         = _norm_campus(campus_txt)
            campus_obj          = campuses.get(campus_code) or Campus.objects.using(using).filter(campus_name=campus_code).first()

            location            = _strip(row[col["location"]])
            day_name            = _as_day_name(_strip(row[col["day_of_week"]]))
            start_time          = _as_time(row[col["start_time"]])
            duration_min = _to_int(row[col["duration"]]) or 0

            if "end_time" in col and pd.notna(row[col["end_time"]]):
                end_time = _to_time(row[col["end_time"]])
            else:
                end_time = _end_time_from_start_and_duration(start_time, duration_min)

            if not end_time:
                raise ValueError("Missing/invalid duration — cannot compute end_time")
            weeks_raw           = _strip(row[col["weeks"]])
            weeks_canon         = _canon_weeks_str(weeks_raw)
            weeks_count         = _count_weeks(weeks_canon)
            staff               = _strip(row[col["staff"]]) if "staff" in col else ""
            size                = int(row[col["size"]]) if ("size" in col and pd.notna(row[col["size"]])) else 0

            # basic identity checks
            unit_code = (subject_code or "")[:6].upper()
            if not unit_code or not day_name or not start_time:
                raise ValueError("Missing unit_code/day_of_week/start_time")

            # --- Unit ---
            unit, _ = Unit.objects.using(using).get_or_create(
                unit_code=unit_code,
                defaults={"unit_name": subject_description or unit_code, "credits": None},
            )
            if subject_description and unit.unit_name != subject_description:
                unit.unit_name = subject_description
                unit.save(using=using, update_fields=["unit_name"])

            # --- UnitCourse (course_id NOT NULL) ---
            uc, _ = UnitCourse.objects.using(using).get_or_create(
                unit=unit,
                course=course,                # << required, never NULL
                campus=campus_obj,            # may be None if campus missing; allowed by DB
                term=term,
                year=year,
                defaults={"status": "Active"},
            )

            # --- MasterClassTime (template) ---
            mct_lookup = dict(
                subject_code=subject_code,
                activity_group_code=activity_group_code,
                activity_code=activity_code,
                day_of_week=day_name,
                start_time=start_time,
            )
            mct_defaults = dict(
                subject_description=subject_description,
                activity_description=activity_description,
                campus=campus_code,
                location=location,
                weeks=weeks_canon,
                teaching_weeks=weeks_count,
                staff=staff,
                size=size,
                faculty="",
                duration=duration_min,
                buffer=0,
                adjusted_size=size,
                student_count=0,
                constraint_count=0,
                cluster="",
                group="",
                show_on_timetable=True,
                available_for_allocation=True,
                updated_at=timezone.now(),
            )
            mct, _ = MasterClassTime.objects.using(using).update_or_create(**mct_lookup, defaults=mct_defaults)

            # --- TimeTable (sessions used by allocation/clash) ---
            slot_id = _slot_key(uc.pk, day_name, start_time, location)
            if slot_id in seen_slots:
                stats.log(
                    f"Row {i+2}: duplicate timetable slot for {unit_code} "
                    f"{day_name} {start_time} {location or ''} — skipped."
                )
                continue
            seen_slots.add(slot_id)

            lookup = dict(
                unit_course=uc,
                campus=campus_obj,   
                day_of_week=day_name,  
                start_time=start_time,
            )

            defaults = dict(
                room=location or "",
                end_time=end_time,         # already computed from start + duration
                master_class=mct,
                tutor_user=None,           # leave pending for manual assignment later
                start_date=None,
                end_date=None,
                updated_at=timezone.now(),
            )

            TimeTable.objects.using(using).update_or_create(**lookup, defaults=defaults)
            stats.inc()


        except Exception as ex:
            stats.log(f"Row {i+2}: {ex}")

    _finalise_log(log, stats, total, ok_status="MASTER_CLASSES")
    job.rows_ok = stats.ok
    job.rows_error = stats.err
    job.ok = stats.err == 0
    job.finished_at = timezone.now()
    job.save(update_fields=["rows_ok", "rows_error", "ok", "finished_at"])

# ===========================
# Tutorial allocations import
# ===========================

def _resolve_current_unit_course(using: str, unit: Unit) -> Optional[UnitCourse]:
    """
    Pick (or lazily create) a UnitCourse for the given unit so later imports never fail.
    """
    uc = (UnitCourse.objects.using(using)
          .filter(unit=unit)
          .order_by("-year", "-created_at")
          .first())
    if uc:
        return uc
    # lazily create with default course & inferred term/year
    term, year = _infer_term_year_from_db(using)
    default_course = _get_default_course(using)
    return UnitCourse.objects.using(using).create(
        unit=unit, course=default_course, campus=None, term=term, year=year, status="Active"
    )

def import_tutorial_allocations_xlsx(file_like, job, using: str):
    """
    Import Timetable rows (allocations) and attach tutors when possible.
    """
    df = pd.read_excel(file_like)
    col = _validate_headers("tutorial_allocations", df)

    stats = ImportStats()
    log = _create_log(getattr(job.file, "name", "tutorial_allocations.xlsx"), job.created_by)

    total = len(df)
    for i, row in df.iterrows():
        try:
            unit_code = _strip(row[col["unit_code"]])
            day = _as_day_name(_strip(row[col["day_of_week"]]))
            start = _as_time(row[col["start_time"]])
            end = _as_time(row[col["end_time"]])
            room = _strip(row[col["room"]])
            email = _strip(row[col["tutor_email"]]).lower()

            if not unit_code or not day or not start or not end:
                raise ValueError("Missing unit_code/day_of_week/start_time/end_time")

            unit = Unit.objects.using(using).filter(unit_code__iexact=unit_code).first()
            if not unit:
                unit = Unit.objects.filter(unit_code__iexact=unit_code).first()
            if not unit:
                raise ValueError(f"No unit '{unit_code}'")

            uc = _resolve_current_unit_course(using, unit)
            if not uc:
                raise ValueError(f"No UnitCourse found for unit '{unit_code}' in current semester")

            tutor = User.objects.filter(email__iexact=email).first() if email else None

            # Try to find a matching MCT; if not found we still create a Timetable entry
            mct = MasterClassTime.objects.using(using).filter(
                subject_code__iexact=unit_code,
                day_of_week=day,
                start_time=start,
            ).first()

            # Natural key for timetable upsert
            lookup = dict(
                unit_course=uc,
                day_of_week=day,
                start_time=start,
                room=room or "",
            )
            defaults = dict(
                end_time=end,
                end_date=None,
                campus=uc.campus if hasattr(uc, "campus") else None,
                master_class=mct,
                tutor_user=tutor,
                updated_at=timezone.now(),
            )

            TimeTable.objects.using(using).update_or_create(
                **lookup, defaults=defaults
            )
            stats.inc()
        except Exception as ex:
            stats.log(f"Row {i + 2}: {ex}")

    _finalise_log(log, stats, total, ok_status="ALLOCATIONS")
    job.rows_ok = stats.ok
    job.rows_error = stats.err
    job.ok = stats.err == 0
    job.finished_at = timezone.now()
    job.save(update_fields=["rows_ok", "rows_error", "ok", "finished_at"])

# ==================
# TUTOR ALLOCATION
# ==================

# ---------- utility ----------

def _apps_get(app_label: str, model_name: str):
    try:
        return django_apps.get_model(app_label, model_name)
    except LookupError:
        return None

def _to_int(value: Any) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(str(value).strip())
    except Exception:
        return None

def _to_time(x: Any) -> Optional[time]:
    """
    Excel times can arrive as:
      - Python datetime.time (already OK)
      - datetime.datetime with a date (take .time())
      - number (Excel time fraction of a day)
      - string ("13:00", "1:00 PM")
    """
    if x is None or x == "":
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, datetime):
        return x.time()
    # numeric excel time fraction
    if isinstance(x, (int, float)):
        seconds = float(x) * 24 * 60 * 60
        h = int(seconds // 3600)
        m = int((seconds % 3600) // 60)
        s = int(seconds % 60)
        return time(h, m, s)
    s = str(x).strip()
    try:
        for fmt in ("%H:%M", "%H:%M:%S", "%H:%M:%S.%f",
                    "%I:%M %p", "%I:%M:%S %p", "%I:%M:%S.%f %p"):
            try:
                return datetime.strptime(s, fmt).time()
            except Exception:
                pass
    except Exception:
        pass
    return None

# ---------- models via apps registry (avoids hard imports & circulars) ----------

EOIImport   = _apps_get("imports", "EOIImport")
MasterEOI   = _apps_get("eoi", "MasterEOI")

# ---------- workbook readers ----------

REQUIRED_EOI_COLS = {
    "unit_code", "tutor_email", "preference", "campus", "qualifications", "availability"
}

EOI_HEADER_ALIASES = {
    "unit_code": {"unit code", "unit", "code"},
    "tutor_email": {"email address", "email", "tutor email"},
    "preference": {"preference", "pref", "rank"},
    "campus": {"campus"},
    "qualifications": {"qualifications", "skills", "notes"},
    "availability": {"availability", "hours", "available hours"},
}

DAY_NORMALIZE = {
    "mon": "Mon", "monday": "Mon",
    "tue": "Tue", "tues": "Tue", "tuesday": "Tue",
    "wed": "Wed", "wednesday": "Wed",
    "thu": "Thu", "thur": "Thu", "thurs": "Thu", "thursday": "Thu",
    "fri": "Fri", "friday": "Fri",
    "sat": "Sat", "saturday": "Sat",
    "sun": "Sun", "sunday": "Sun",
}

def _best_header_map(header_cells: Iterable[str], aliases: Dict[str, Set[str]]) -> Dict[int, str]:
    """
    Given a row of header strings and a mapping of canonical->aliases,
    return a dict: {column_index: canonical_name}
    """
    header_map: Dict[int, str] = {}
    for idx, raw in enumerate(header_cells):
        label = _strip(raw).lower()
        for canonical, alias_set in aliases.items():
            if label == canonical:
                header_map[idx] = canonical
                break
            if label in alias_set:
                header_map[idx] = canonical
                break
    return header_map

# ---------- reference data upserts ----------

def _ensure_campuses(rows: Iterable[Dict[str, Any]], *, using: str) -> int:
    names = {_strip(r.get("campus")) for r in rows if _strip(r.get("campus"))}
    if not names:
        return 0

    # normalise to choice codes if needed
    codes = {_CAMPUS_ALIAS.get(n.lower(), n) for n in names}

    existing = set(
        Campus.objects.using(using)
        .filter(campus_name__in=codes)
        .values_list("campus_name", flat=True)
    )

    to_create = [Campus(campus_name=code, campus_location=code) for code in (codes - existing)]
    if to_create:
        Campus.objects.using(using).bulk_create(to_create, ignore_conflicts=True)
    return len(to_create)

def _ensure_units(rows: Iterable[Dict[str, Any]], *, using: str) -> int:
    codes = {_strip(r.get("unit_code")) for r in rows if _strip(r.get("unit_code"))}
    if not codes:
        return 0
    existing = set(
        Unit.objects.using(using).filter(unit_code__in=codes).values_list("unit_code", flat=True)
    )
    to_create = [Unit(unit_code=c, unit_name=c) for c in (codes - existing)]
    if to_create:
        Unit.objects.using(using).bulk_create(to_create, ignore_conflicts=True)
    return len(to_create)


# ---------- writers ----------

def _write_eoi_staging(rows: list[dict], *, job, using: str) -> int:
    """Keep your current staging log behaviour (eoi_imports)."""
    if EOIImport is None:
        # staging table not defined; skip quietly
        return 0
    objs = [
        EOIImport(
            job=job,
            unit_code=r["unit_code"],
            tutor_email=r["tutor_email"],
            preference=r["preference"],
            campus=r["campus"],
            qualifications=r["qualifications"],
            availability=r["availability"],
        )
        for r in rows
    ]
    if objs:
        EOIImport.objects.using(using).bulk_create(objs)
    return len(objs)

def _upsert_master_eoi(rows: List[Dict[str, Any]], *, using: str) -> Tuple[int, int]:
    """
    Upsert into eoi.MasterEOI by (unit_code, tutor_email).
    Returns (created, updated).
    """
    if MasterEOI is None:
        return 0, 0
    # First, create any missing pairs
    pairs = {(r["unit_code"], r["tutor_email"]) for r in rows}
    existing_pairs = set(
        MasterEOI.objects.using(using)
        .filter(unit_code__in=[p[0] for p in pairs], tutor_email__in=[p[1] for p in pairs])
        .values_list("unit_code", "tutor_email")
    )
    to_create = []
    for r in rows:
        key = (r["unit_code"], r["tutor_email"])
        if key not in existing_pairs:
            to_create.append(
                MasterEOI(
                    unit_code=r["unit_code"],
                    tutor_email=r["tutor_email"],
                    preference=r["preference"],
                    campus=r["campus"],
                    qualifications=r["qualifications"],
                    availability=r["availability"],
                )
            )
    created = 0
    if to_create:
        MasterEOI.objects.using(using).bulk_create(to_create, ignore_conflicts=True)
        created = len(to_create)

    # Update fields on existing rows
    # (do a second pass to pick only the rows that already exist)
    to_update = []
    existing_qs = MasterEOI.objects.using(using).filter(
        unit_code__in=[p[0] for p in existing_pairs],
        tutor_email__in=[p[1] for p in existing_pairs],
    )
    existing_map: Dict[Tuple[str, str], Any] = {
        (m.unit_code, m.tutor_email): m for m in existing_qs
    }
    for r in rows:
        key = (r["unit_code"], r["tutor_email"])
        m = existing_map.get(key)
        if not m:
            continue
        dirty = False
        if m.preference != (r["preference"] or 0):
            m.preference = r["preference"] or 0
            dirty = True
        if m.campus != r["campus"]:
            m.campus = r["campus"]
            dirty = True
        if m.qualifications != r["qualifications"]:
            m.qualifications = r["qualifications"]
            dirty = True
        if m.availability != (r["availability"] or 0):
            m.availability = r["availability"] or 0
            dirty = True
        if dirty:
            to_update.append(m)

    updated = 0
    if to_update:
        MasterEOI.objects.using(using).bulk_update(
            to_update, ["preference", "campus", "qualifications", "availability"]
        )
        updated = len(to_update)

    return created, updated


def _write_master_classes(rows: List[Dict[str, Any]], *, using: str) -> int:
    """
    Insert all class rows into MasterClassTime.
    We don't try to deduplicate here (coordinators may re-upload new terms);
    if you want deduplication, add a unique_together in the model.
    """
    objs = [
        MasterClassTime(
            unit_code=r["unit_code"],
            campus=r["campus"],
            activity=r["activity"],
            class_group=r["group"],
            day=r["day"],
            start_time=r["start_time"],
            end_time=r["end_time"],
        )
        for r in rows
    ]
    if objs:
        MasterClassTime.objects.using(using).bulk_create(objs)
    return len(objs)


# ---------- public import API (used by the view) ----------

def _parse_casual_master_eoi(file_like) -> list[dict]:
    """
    Parse the one-sheet 'Casual Master EOI Spreadsheet.xlsx' and
    return EOI-row dicts compatible with import_eoi_excel().
    One EOI row per (tutor_email, unit_code).
    """
    df = pd.read_excel(file_like)  # first sheet

    # normalise headers -> actual column names
    cols = {str(c).strip(): str(c).strip() for c in df.columns}
    lower = {c.lower(): c for c in cols}

    def get_col(*aliases):
        for a in aliases:
            a_l = a.lower()
            if a_l in lower:
                return lower[a_l]
            # allow substring contains for very long labels
            for k in lower:
                if a_l in k:
                    return lower[k]
        return None

    c_name        = get_col("Name")
    c_email       = get_col("Email Address", "Email", "Email Address*")
    c_you_are     = get_col("You are")
    c_location    = get_col("Tutoring Location", "Location", "Campus")
    c_hours       = get_col("Total number of tutoring hours you wish to work")
    c_scholarship = get_col("Do you receive a Scholarship")
    c_gpa         = get_col("What is your GPA", "GPA")
    c_supervisor = get_col(
        "Please indicate your supervisor name / references in School of ICT",
        "Please indicate your supervisor name / references",
        "Please indicate your supervisor name",
        "reference name", "references", "supervisor"
    )
    c_applied     = get_col("Please select up to five units", "applied unit")
    c_skills      = get_col("What technical and/or other skills", "Why do you want to teach this unit")
    c_experience  = get_col("Have you tutored any of the ICT units")
    c_transcript  = get_col("upload your transcript", "transcript")
    c_cv          = get_col("upload your CV", "cv")

    out: list[dict] = []
    for _, row in df.iterrows():
        email = str(row.get(c_email) or "").strip().lower()
        if not email or "@" not in email:
            continue  # skip blank/invalid

        # unit codes from the “select up to five units” cell
        applied_raw = str(row.get(c_applied) or "")
        unit_codes = list({u.upper() for u in UNIT_CODE_RX.findall(applied_raw.upper())})
        if not unit_codes:
            # still create a placeholder row so coordinator can set preference later if needed
            unit_codes = []

        tutor_name = str(row.get(c_name) or "").strip()
        you_are    = str(row.get(c_you_are) or "").strip()
        location   = str(row.get(c_location) or "").strip()
        skills     = str(row.get(c_skills) or "").strip()
        experience = str(row.get(c_experience) or "").strip()
        hours      = str(row.get(c_hours) or "").strip()
        scholarship= str(row.get(c_scholarship) or "").strip()
        gpa        = str(row.get(c_gpa) or "").strip()
        supervisor = str(row.get(c_supervisor) or "").strip()
        transcript = str(row.get(c_transcript) or "").strip()
        cv         = str(row.get(c_cv) or "").strip()

        extras = {
            "tutor_name": tutor_name,
            "eoi_status_text": you_are,
            "location_text": location,
            "gpa": gpa,
            "supervisor": supervisor,
            "applied_units": unit_codes,
            "tutoring_experience": experience,
            "hours_available": hours,
            "scholarship_received": scholarship,
            "transcript_link": transcript,
            "cv_link": cv,
        }

        # emit one row per applied unit (or a single row with empty unit_code if none)
        if unit_codes:
            for u in unit_codes:
                out.append({
                    "unit_code": u,
                    "tutor_email": email,
                    "preference": 0,
                    "campus": location,
                    "qualifications": skills,
                    "availability": hours,
                    "_extra": extras,
                })
        else:
            out.append({
                "unit_code": "",
                "tutor_email": email,
                "preference": 0,
                "campus": location,
                "qualifications": skills,
                "availability": hours,
                "_extra": extras,
            })

    if not out:
        raise ValidationError("No valid EOI rows found in the Casual Master EOI spreadsheet.")
    return out

def import_eoi_excel(fileobj, job, using: str):
    """
    Parse the Casual Master EOI Spreadsheet and write rows directly into eoi.EoiApp,
    creating Units/Campuses/Users as needed in the current semester DB.
    """
    rows = _parse_casual_master_eoi(fileobj)

    created = 0
    updated = 0
    with transaction.atomic(using=using):
        for r in rows:
            unit_code = (r.get("unit_code") or "").strip().upper()
            if not unit_code:
                continue
            # ensure unit
            unit, _ = Unit.objects.using(using).get_or_create(
                unit_code=unit_code,
                defaults={"unit_name": unit_code}
            )
            # campus
            campus_txt = (r.get("campus") or "").strip()
            campus_key = CAMPUS_MAP.get(campus_txt.lower(), None)
            campus = None
            if campus_key:
                campus, _ = Campus.objects.using(using).get_or_create(
                    campus_name=campus_key,
                    defaults={"campus_location": campus_txt or campus_key}
                )
            # user
            email = (r.get("tutor_email") or "").strip().lower()
            if not email or "@" not in email:
                continue
            # Try to split name from extras
            tutor_name = ""
            extras = r.get("_extra") or {}
            if extras:
                tutor_name = (extras.get("tutor_name") or "").strip()
            first, last = "", ""
            if tutor_name:
                parts = tutor_name.split()
                first = parts[0]
                last = " ".join(parts[1:]) if len(parts) > 1 else ""
            user, _ = User.objects.using(using).get_or_create(
                email=email,
                defaults={"first_name": first, "last_name": last, "is_active": False}
            )
            # upsert EoiApp (SCD handled by model .save())
            obj, created_flag = EoiApp.objects.using(using).get_or_create(
                applicant_user=user,
                unit=unit,
                campus=campus,
                defaults={
                    "status": "Submitted",
                    "remarks": "",
                    "preference": int(r.get("preference") or 0),
                    "qualifications": r.get("qualifications") or "",
                    "availability": r.get("availability") or "",
                    "tutor_email": email,
                    "tutor_name": tutor_name,
                    "tutor_current": extras.get("eoi_status_text") or "",
                    "location_text": extras.get("location_text") or campus_txt or "",
                    "gpa": (float(extras.get("gpa")) if str(extras.get("gpa") or "").strip() not in {"", "nan"} else None),
                    "supervisor": extras.get("supervisor") or "",
                    "applied_units": extras.get("applied_units") or None,
                    "tutoring_experience": extras.get("tutoring_experience") or "",
                    "hours_available": (int(extras.get("hours_available")) if str(extras.get("hours_available") or "").strip().isdigit() else None),
                    "scholarship_received": None if (extras.get("scholarship_received") is None) else str(extras.get("scholarship_received")).strip().lower() in {"y", "yes", "true", "1"},
                    "transcript_link": extras.get("transcript_link") or "",
                    "cv_link": extras.get("cv_link") or "",
                }
            )
            if not created_flag:
                # update mutable fields and save to trigger SCD if changed
                obj.preference = int(r.get("preference") or obj.preference or 0)
                obj.qualifications = r.get("qualifications") or obj.qualifications or ""
                obj.availability = r.get("availability") or obj.availability or ""
                # extras
                if tutor_name:
                    obj.tutor_name = tutor_name
                ex = extras
                if ex:
                    if ex.get("eoi_status_text"): obj.tutor_current = ex["eoi_status_text"]
                    if ex.get("location_text"): obj.location_text = ex["location_text"]
                    try:
                        g = ex.get("gpa"); obj.gpa = float(g) if str(g or "").strip() not in {"", "nan"} else obj.gpa
                    except Exception: pass
                    sup = ex.get("supervisor")
                    if sup is not None:
                        s = str(sup).strip()
                        obj.supervisor = "" if s.lower() in {"", "nan"} else s
                    if ex.get("applied_units"): obj.applied_units = ex["applied_units"]
                    if ex.get("tutoring_experience"): obj.tutoring_experience = ex["tutoring_experience"]
                    try:
                        h = ex.get("hours_available"); obj.hours_available = int(h) if str(h or "").strip().isdigit() else obj.hours_available
                    except Exception: pass
                    if ex.get("scholarship_received") is not None:
                        v = str(ex["scholarship_received"]).strip().lower()
                        obj.scholarship_received = True if v in {"y","yes","true","1"} else False if v in {"n","no","false","0"} else obj.scholarship_received
                    if ex.get("transcript_link"): obj.transcript_link = ex["transcript_link"]
                    if ex.get("cv_link"): obj.cv_link = ex["cv_link"]
                obj.save()
                updated += 1
            else:
                created += 1

    return {
        "result": "ok",
        "target": "eoi_app",
        "inserted": created,
        "updated": updated,
    }

# =================
# Import dispatcher
# =================

IMPORT_DISPATCH = {
    "eoi": import_eoi_excel,
    "master_class_list": import_master_classes_xlsx,
    "master_classes": import_master_classes_xlsx,
    "tutorial_allocations": import_tutorial_allocations_xlsx,
}

